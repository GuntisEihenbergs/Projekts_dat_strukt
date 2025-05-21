from openpyxl import Workbook, load_workbook 
import requests
from bs4 import BeautifulSoup
adrese = "https://www.deltaco.lv/lv/aksesuari/gaming/mouses.html"
aktīvs = requests.get(adrese)
both = []
if aktīvs.status_code == 200:
    saturs = BeautifulSoup(aktīvs.content, 'html.parser')
    for label in saturs.find_all(class_="price-box"):
        old = label.find(class_="price old-5")
        discount= label.find(class_="special-price")
        if discount:
            new = discount.find(class_="price")
        if old and new:
            both.append((old.get_text(strip=True), new.get_text(strip=True)))
else:
    print("Kļūda: ", aktīvs.status_code)
wb = load_workbook('kopa/Book5.xlsx')
ws=wb["Sheet1"]
wb.remove(ws)
if 'First' in wb.sheetnames:
    ws = wb['First']
else:
    ws = wb.create_sheet('First')
for x,(one,two) in enumerate(both,start=1):
    ws[f'A{x}']=one
    ws[f'B{x}']=two
    oldp = float(one.replace('€','').replace(',','.'))
    newp = float(two.replace('€','').replace(',','.'))
    izm = oldp-newp
    ws[f'C{x}']=f"{izm:.2f}".replace('.',',')+'€'
wb.save('kopa/Book5.xlsx')
wb.close()